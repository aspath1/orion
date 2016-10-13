#!/usr/bin/env python
# coding: utf-8
# Version 1 - 2016/10/10
# Creator - David Varnum
'''
This script will add multiple nodes to SolarWinds Orion.

1. Update the 'add_nodes.xlsx' spreadsheet with the node values and properties.
2. Execute this script, providing your username and password.
3. The script will iterate through the spreadsheet, adding each node to Orion.
	- Adds node, pollers, custom properties, and NCM values
'''

from __future__ import print_function
from orionsdk import SwisClient
import re
import requests
import getpass
from pprint import pprint as pp
from openpyxl import load_workbook

class NodeManager:
	def __init__(self, username, password):
		
		# Change 'orion' to your Orion NPM server name or IP.
		self._swis = SwisClient('orion', username, password)

	def add_node(self, **node):	
		self._node_element(**node)
		self._node_pollers()
		self._node_custom_props(**node)
		self._poll()
		self._node_ncm(**node)

	def _node_element(self, **node):	
		
		# Extract IP Address and node name from node variables
		ip_address = node['IP_Address']
		node_name = node['Caption']
		
		# Setup properties for the new node
		props = {
			'IPAddress': ip_address,
			'EngineID': 5,
			'ObjectSubType': 'SNMP',
			'SNMPVersion': 3,
			'Caption': node_name,
			'SNMPV3AuthKey': 'YourKey',  # Enter your key here
			'SNMPv3AuthKeyIsPwd': True,
			'SNMPv3AuthMethod': 'SHA1',
			'SNMPv3PrivKey': 'YourKey',  # Enter your key here
			'SNMPv3PrivKeyIsPwd': True,
			'SNMPv3PrivMethod': 'AES128',
			'SNMPV3Username': 'YourUser'  # Enter your SNMPv3 Username here
		}
		
		# Create the node
		print("Adding node {}... ".format(props['IPAddress']), end="")
		self._results = self._swis.create('Orion.Nodes', **props)
		print("DONE!")
		
		# Extract nodeID from the results
		self._nodeid = self._parse_node(self._results)
	
	def _parse_node(self, results):
		return re.search('(\d+)$', self._results).group(0)

	def _node_pollers(self):
		
		# Setup poller status for the node
		pollers_enabled = {
			'N.Status.ICMP.Native': True,
			'N.Status.SNMP.Native': False,
			'N.ResponseTime.ICMP.Native': True,
			'N.ResponseTime.SNMP.Native': False,
			'N.Details.SNMP.Generic': True,
			'N.Uptime.SNMP.Generic': True,
			'N.Cpu.SNMP.HrProcessorLoad': True,
			'N.Memory.SNMP.NetSnmpReal': True,
			'N.AssetInventory.Snmp.Generic': True,
			'N.Topology_Layer3.SNMP.ipNetToMedia': False,
			'N.Routing.SNMP.Ipv4CidrRoutingTable': False
		}
	
		pollers = []
		for k in pollers_enabled:
			pollers.append(
				{
					'PollerType': k,
					'NetObject': 'N:' + self._nodeid,
					'NetObjectType': 'N',
					'NetObjectID': self._nodeid,
					'Enabled': pollers_enabled[k]
				}
			)
	
		# Add node to pollers
		for poller in pollers:
			print("  Adding poller type: {} with status {}... ".format(poller['PollerType'], poller['Enabled']), end="")
			response = self._swis.create('Orion.Pollers', **poller)
			print("DONE!")
		
	def _node_custom_props(self, **node):

		# Copy 'node' dict to new dict 'props' and remove keys not relative to custom properties
		props = node
		ignore_keys = ('IPAddress', 'Caption', 'NodeGroup', 'DeviceTemplate', 'ConnectionProfile')
		for k in ignore_keys:
			props.pop(k, None)

		# Add custom properties to node
		for k,v in props.items():
			print("  Adding custom property: {} with value {}... ".format(k, v), end="")
			self._swis.update(self._results + '/CustomProperties', **{k: v})
			print("DONE!")

	def _poll(self):
		
		# Poll the node
		print("  Polling node... ", end="")
		self._swis.invoke('Orion.Nodes', 'PollNow', 'N:' + self._nodeid)
		print("DONE!")

	def _node_ncm(self, **node):  
	    
	    # Add node to NCM
	    self._swis.invoke('Cirrus.Nodes', 'AddNodeToNCM', self._nodeid)

	    # Lookup the NCM NodeID, which is a Guid  
	    ncmNodeID = self._swis.query('SELECT NodeID FROM Cirrus.Nodes WHERE CoreNodeID=@node', node=self._nodeid)['results'][0]['NodeID']  
	  
	    # Fetch the NCM Node object  
	    ncmNode = self._swis.invoke('Cirrus.Nodes', 'GetNode', ncmNodeID)  
	  
	    # Modify our local copy of the NCM Node object  
	    ncmNode['ConnectionProfile'] = node['ConnectionProfile']
	    ncmNode['DeviceTemplate'] = node['DeviceTemplate'] 
	    ncmNode['NodeGroup'] = node['NodeGroup']
	  
	    # Commit our changes  
	    print("  Adding node to NCM... ", end="")
	    self._swis.invoke('Cirrus.Nodes', 'UpdateNode', ncmNode)
	    print("DONE!")

def xlsx_dict_list(variables_file, sheetname):

	wbook = load_workbook(variables_file, sheetname)
	wsheet = wbook[sheetname]
	
	header = [cell.value for cell in wsheet[1]]
	
	dict_list = []
	
	for row in wsheet.iter_rows(min_row=2):
		values = {}
		for key, cell in zip(header, row):
		    values[key] = cell.value
		dict_list.append(values)
	return dict_list

def main():
	
	username = raw_input('Username: ')
	password = getpass.getpass()

	# Initialize the NodeManager class
	nm = NodeManager(username, password)

	# Disable SSL warning messages
	requests.packages.urllib3.disable_warnings()

	# Fetch key/value pairs for each row in the spreadsheet
	all_nodes = xlsx_dict_list('add_nodes.xlsx', 'Sheet1')

	# Iterate over list of nodes, invoking the add_node method within the NodeManager class
	for node in all_nodes:
		nm.add_node(**node)

if __name__ == '__main__':
    main()